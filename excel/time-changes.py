#!/usr/bin/env python3

from argparse import ArgumentParser
from openpyxl import load_workbook, utils
import warnings


def createFullCoordinate(cell, sheet):
    return f"{utils.quote_sheetname(sheet.title)}!{cell.coordinate}"


def calculateVariableDifference(relevantVariableCells, dataRow, variableName, firstTimeSpot, lastTimeSpot, targetCellAbsolute):
    firstHeaderCell = list(filter(lambda x: str.startswith(x.value, firstTimeSpot + "_")
                                  and variableName in x.value, relevantVariableCells))[0]
    lastHeaderCell = list(filter(lambda x: str.startswith(x.value, lastTimeSpot + "_")
                                 and variableName in x.value, relevantVariableCells))[0]
    firstDataCell = list(dataRow)[firstHeaderCell.column-1]
    lastDataCell = list(dataRow)[lastHeaderCell.column-1]
    absoluteDifferenceFormular = f"{createFullCoordinate(lastDataCell, readingWorksheet)}-{createFullCoordinate(firstDataCell, readingWorksheet)}"
    absolutDifference = f"={absoluteDifferenceFormular}"
    relativeDifference = f"=({absoluteDifferenceFormular})/{createFullCoordinate(firstDataCell, readingWorksheet)}"
    return (absolutDifference, relativeDifference, firstDataCell.style)


# arguments of the script
parser = ArgumentParser()
parser.add_argument("-f", "--file", dest="filename", help="file to read",
                    metavar="FILENAME", default="testfile.xlsx", required=True)
parser.add_argument("-s", "--sheetname", dest="sheetname", help="sheet to read", metavar="SHEETNAME")
parser.add_argument("-v", "--values", dest="values", required=True,
                    help="Which values to generate. Possible options are 'absolute' or 'relative'", metavar="VALUES")
args = parser.parse_args()
sheetname = args.sheetname
filename = args.filename
# TODO: REFACTOR TO ALLOW ONLY THOSE TWO VALUES
valuesType = args.values
isAbsolute = valuesType == 'absolute'
timespots = ['Pre', 'Mid', 'Post']

# get excel sheet to read from
workbook = load_workbook(filename)
readingWorksheet = workbook[workbook.sheetnames[0]] if sheetname == None else workbook[sheetname]

# create excel sheet to write to
newWorksheetName = f"{valuesType}-changes-over-time"
# remove worksheet, if it already exists
if newWorksheetName in workbook.sheetnames:
    workbook.remove(workbook[newWorksheetName])
# create sheet from scratch
writingWorksheet = workbook.create_sheet(newWorksheetName)

# get required data
rowList = list(readingWorksheet.rows)

# get all variables
# get all items in the top row
allVariableCells = list(rowList[0])
# remove empty variables
notEmptyVariableCells = list(filter(lambda x: x.value != None, allVariableCells))
# get variables for all required time spots (--> maybe there are additional variables which we do not care about)
relevantVariableCells = list(filter(lambda x: any(str.startswith(x.value, timespot + "_")
                                                  for timespot in timespots), notEmptyVariableCells))
# get distinct variable names (without time spot)
allVariableNames = list(map(lambda x: x.value.partition('_')[-1], relevantVariableCells))
distinctVariableNames = list(dict.fromkeys(allVariableNames))
# remove variable names which do not have the appropriate amount of values (--> like amount of time spots)
relevantVariableNames = list(filter(lambda x: allVariableNames.count(x) == len(timespots), distinctVariableNames))

# catch variable names which are not equal for all time spots
irrelevantVariableNames = [x for x in allVariableNames if x not in relevantVariableNames]
if len(irrelevantVariableNames) > 0:
    irrelevantVariableNames.sort()
    warnings.warn('The following variables have different names amongst the different time spots:\n\n' +
                  str(irrelevantVariableNames))

# go over all relevant variables and calculate differences between first and last timespot
# then write those differences to the excel file
numberOfUsedColumns = len(allVariableCells)
firstTimeSpot = timespots[0]
lastTimeSpot = timespots[2]
for (variableIndex, variableName) in enumerate(relevantVariableNames):
    # name the headers of the columns appropriately
    firstColumnNumber = 2 + variableIndex
    headerCell = writingWorksheet.cell(row=1, column=firstColumnNumber)
    headerCell.value = f"{valuesType}-{firstTimeSpot}-{lastTimeSpot}-{variableName}"
    headerCell.style = 'Headline 3'

    # go over all rows, calculate differences between values for the specific variable
    # and finally write the difference values to extra columns, whose headers were already written
    for (rowIndex, row) in enumerate(rowList[1:]):
        dataCell = writingWorksheet.cell(row=rowIndex + 2, column=firstColumnNumber)
        difference = calculateVariableDifference(
            relevantVariableCells, row, variableName, firstTimeSpot, lastTimeSpot, dataCell)
        dataCell.value = difference[0] if isAbsolute else difference[1]
        dataCell.style = difference[2]


workbook.save(filename)
