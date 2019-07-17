#!/usr/bin/env python3

from argparse import ArgumentParser
from openpyxl import load_workbook
from openpyxl.chart import (LineChart, Reference)

# arguments of the script
parser = ArgumentParser()
parser.add_argument("-f", "--file", dest="filename",
                    help="file to read", metavar="FILENAME", default="testfile.xlsx")
parser.add_argument("-s", "--sheetname", dest="sheetname",
                    help="sheet to read", metavar="SHEETNAME")
parser.add_argument("-t", "--targetsheet", dest="targetsheet",
                    help="sheet to write the result to", metavar="TARGETSHEET")
parser.add_argument("-r", "--sourcerownumber", dest="sourcerow", required=True,
                    help="row number to use for the chart", metavar="SOURCEROWNUMBER")
args = parser.parse_args()
sheetname = args.sheetname
filename = args.filename
targetsheet = args.targetsheet
sourceRowNumber = int(args.sourcerow)

# get excel sheet to read from
workbook = load_workbook(filename)
readingWorksheet = workbook[workbook.sheetnames[0]] if sheetname == None else workbook[sheetname]

# create excel sheet to write to
newWorksheetName = "changes-as-chart" if targetsheet == None else targetsheet
# remove worksheet, if it already exists
if newWorksheetName in workbook.sheetnames:
    workbook.remove(workbook[newWorksheetName])
# create sheet from scratch
writingWorksheet = workbook.create_sheet(newWorksheetName)

# get required data
rowList = list(readingWorksheet.rows)
headerRow = rowList[0][1:]
dataRow = rowList[sourceRowNumber-1][1:]
columnNumber = len(headerRow)

# write row headers
writingWorksheet.cell(1, 1, 'Point in Time')
writingWorksheet.cell(2, 1, 'Pre')
writingWorksheet.cell(3, 1, 'Post')

# write values for chart to a sheet, so that table can reference these values
for (index, column) in enumerate(headerRow):
    writingWorksheet.cell(1, index+2, column.value)
    writingWorksheet.cell(2, index+2, 0)
    writingWorksheet.cell(3, index+2, dataRow[index].value)


c1 = LineChart()
c1.title = "Relative Changes as chart"
c1.style = 13
c1.y_axis.title = "Relative Change"
c1.x_axis.title = "Time"
data = Reference(writingWorksheet, min_col=2, min_row=1, max_col=columnNumber+1, max_row=3)
c1.add_data(data, titles_from_data=True)


# color each line within the chart looking at the value
for (index, column) in enumerate(headerRow):
    line = c1.series[index]
    line.graphicalProperties.line.width = 100
    color = column.fill.start_color.index
    line.graphicalProperties.line.solidFill = color

writingWorksheet.add_chart(c1, "A5")

workbook.save(filename)
